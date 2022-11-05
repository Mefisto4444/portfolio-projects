from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.common import exceptions
from sys import argv
import openpyxl
from openpyxl.styles.fonts import Font

class Scraper():
    def __init__(self) -> None:
        self.options = webdriver.ChromeOptions()
        self.options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.options.headless = True
        self.options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.5112.102 Safari/537.36 OPR/90.0.4480.117')
        self.driver = Service("C:\Program Files (x86)\chromedriver.exe")
        self.browser = webdriver.Chrome(service=self.driver,options=self.options)

    def connect(self, url:str):
        self.browser.get(url=url)
        self.browser.get(url=url)
        return self.browser


def main_gen(urls:list):
    s = Scraper()
    for url in urls:
        yield s.connect(url=url)


def handler(find_element:webdriver.Chrome, by:By, value:str):
    try:
        return find_element.find_element(by=by,value=value).text
    except exceptions.NoSuchElementException:
        return ""

def elements_handler(find_elements:webdriver.Chrome, by:By, value:str):
    try:
        for element in find_elements.find_elements(by=by, value=value):
            yield element.text
    except exceptions.NoSuchElementException:
        yield ""


def main(connection_generator):
    for connection in connection_generator:
        for element in connection.find_elements(By.TAG_NAME,"article"):
            job_title = element.find_element(By.CSS_SELECTOR,"[data-automation='jobTitle']").text
            company = handler(element,By.CSS_SELECTOR,"[data-automation='jobCompany']")
            location = ",".join(list(elements_handler(element,By.CSS_SELECTOR,"[data-automation='jobLocation']")))
            classification = element.find_element(By.CSS_SELECTOR,"[data-automation='jobClassification']").text
            subclassification = element.find_element(By.CSS_SELECTOR,"[data-automation='jobSubClassification']").text
            offers = ",".join([offer.find_element(By.CSS_SELECTOR,"span").text for offer in element.find_elements(By.TAG_NAME,"li")])
            description = element.find_element(By.CSS_SELECTOR,"[data-automation='jobShortDescription']").text
            premium = handler(element, By.CSS_SELECTOR,"[data-automation='jobPremium']")
            date = handler(element, By.CSS_SELECTOR,"[data-automation='jobListingDate']")
            price = handler(element,By.CSS_SELECTOR,"[data-automation='jobSalary']")
            yield [job_title, company, price, location, date, premium, description, offers, classification, subclassification]
    connection.quit()

def speadsheet_creator(seek_data:list, name="seek_scraper.xlsx"):
    wb = openpyxl.Workbook()
    try:
        del wb["Sheet"]
    except:
        pass
    wb.create_sheet('seek_data')
    headers = ['job title', 'company', 'price', 'location', 'date', 'premium', 'description', 'offers', 'classification', 'subclassification']
    for i in range(len(headers)):
        wb['seek_data'].cell(1,i+1).font = Font(name='Calibri',bold=True)
        wb['seek_data'].cell(1,i+1).value = headers[i]
    wb['seek_data'].freeze_panes = "L2"

    for row in range(len(seek_data)):
        for col in range(len(headers)):
            wb["seek_data"].cell(row=row+2,column=col+1).value = f"{seek_data[row][col]}"
    wb.save(name)
    

if __name__ == '__main__':
    urls = list(map(lambda url:url+"?page=1", argv[1].split(" ")))
    url_types = urls.copy()
    if len(argv) == 3:
        page_num = int(argv[2])
        urls = list(map(lambda url:[url],urls))
        for i in range(2,page_num+1):
            for url_list in urls:
                url_list.append(f"{url_list[0].replace('page=1',f'page={i}')}")
    urls = [url for url_list in urls for url in url_list]      

    #urls = ["https://www.seek.com.au/jobs-in-trades-services?page=1", "https://www.seek.com.au/jobs-in-accounting?page=1"]
    connection_gen = main_gen(urls=urls)
    seek_data = list(main(connection_generator=connection_gen))
    speadsheet_creator(seek_data=seek_data)


