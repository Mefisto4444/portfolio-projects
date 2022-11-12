import requests, bs4, openpyxl
from openpyxl.styles import Font
class SenatorScraper:

    def __init__(self, urls:list) -> None:
        self.urls = urls

    def connect(self, url:str):
        res = requests.get(url=url)
        try:
            res.raise_for_status()
        except Exception as Ex:
            return f"Error {Ex}"
        return res

    def senator_info_generator(self,senators:bs4.ResultSet[bs4.Tag]):
        for senator in senators:
            name = senator.select_one('h4').text
            link = "https://www.nysenate.gov" + senator.select_one('a').get("href")
            senator_soup = bs4.BeautifulSoup(s.connect(link).text,'html.parser')
            company = senator_soup.select_one("p[class='c-senator-hero--roles']").text.strip("\n\t\r\s ")
            senator_c_soup = bs4.BeautifulSoup(s.connect(link+"/contact").text,'html.parser')
            for vcard in senator_c_soup.select('div[class="location vcard"]'):
                if vcard.select_one('span[class="fn"]') == None:
                    continue
                if vcard.select_one('span[class="fn"]').text.strip() != "Albany Office":
                    continue
                else:
                    break
            contact_info_soup = vcard
            telephone = contact_info_soup.select_one('div[itemprop="telephone"]').text.split(" ")[-1]
            email = senator_c_soup.select_one("div[class='c-block--senator-email']").select_one('a').text
            yield name, company, telephone, email, link

def excel_output(senator_data:list|tuple, name="SenatorScraper.xlsx"):
    wb = openpyxl.Workbook()
    try:
        del wb["Sheet"]
    except:
        pass
    sheet = wb.create_sheet('senators')
    headers = ['name', 'company', 'telephone', 'email', 'link']
    for i in range(len(headers)):
        sheet.cell(row=1,column=i+1).font = Font('Calibri',bold=True)
        sheet.cell(row=1,column=i+1).value = headers[i]
        for row in range(len(senator_data)):
            sheet.cell(row=row+2,column=i+1).value = senator_data[row][i]
    sheet.freeze_panes = "F2"


    wb.save(name)

if __name__ == '__main__':
    
    s = SenatorScraper(["https://www.nysenate.gov/senators-committees"])

    soup = bs4.BeautifulSoup(s.connect(s.urls[0]).text,'html.parser')

    senators = soup.select('div[class="view-content"] > div')
    excel_output(list(s.senator_info_generator(senators=senators)))

        
