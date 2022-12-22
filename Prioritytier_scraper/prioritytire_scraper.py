import asyncio, time, openpyxl
from openpyxl.styles import Font
from playwright.async_api import async_playwright, Playwright
from selectolax.parser import HTMLParser

async def run(playwright:Playwright, urls):
    print("Przygotowywanie dokumentu......")
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('prioritytire')
    sheet.freeze_panes = "K2"
    col_names = ['marka', 'model', 'parametry', 'cechy', 'opinia',  'ilość_opinii', 'cena_bez_podatku', 'cena', 'ilość_w_magazynie']
    for i in range(len(col_names)):
        sheet.cell(row=1, column=i+1).font = Font(bold=True)
        sheet.cell(row=1, column=i+1).value = col_names[i]
    #Kolumny: marka, model, parametry, cechy, opinia, ilośc opinii, cena bez podatku, cena, ilość w magazynie
    print("Tworzenie obiektu browser....")
    chromium = playwright.chromium # or "firefox" or "webkit".
    browser = await chromium.launch(headless=True)
    page = await browser.new_page()
    last_row = 2
    print("Pełzanie po stronach.....")
    for url in urls:
        print('last_row:',last_row,'-- pętla stron')
        print(f"Pełzanie po {page.url}")
        await page.goto(url)
        await page.wait_for_selector('article[class="card"]')
        page_code = await page.inner_html('body')
        tree = HTMLParser(page_code)
        for card in tree.css('article[class="card"]'):
            row = last_row + 1
            print('row', row, '--pętla ofert', page.url)
            features = ""
            sheet.cell(row=row, column=1).value = card.css_first('p[data-test-info-type="brandName"]').text()
            sheet.cell(row=row, column=2).value =  card.css_first('h4[class="card-title"] > a').text()
            sheet.cell(row=row, column=3).value =  card.css_first('span[class="product-sku ng-binding"]').text()
            for feature in card.css_first('ul[class="ppIcons"]').css('span[class="ng-binding"]'):
                features = features + feature.text().strip() + "," 
            sheet.cell(row=row, column=4).value = features
            try:
                sheet.cell(row=row, column=5).value =  float(tree.css_first('span[class="sr-only"]').text().split(" ")[0])
                sheet.cell(row=row, column=6).value = int(card.css_first('a[class="text-m"]').text().split(" ")[0])
            except:
                sheet.cell(row=row, column=5).value = "NULL"
                sheet.cell(row=row, column=6).value = "NULL"
            sheet.cell(row=row, column=7).value = card.css_first('span[class="price price--withoutTax ng-binding"]').text().lstrip("$")
            try:
                sheet.cell(row=row, column=8).value = card.css_first('span[class="price price--rrp ng-binding"]').text().lstrip("$")
            except:
                sheet.cell(row=row, column=8).value = "NULL"
            sheet.cell(row=row, column=9).value = card.css_first('div[class="LoginClubMSGOuter"]').css('div')[-1].text()
            last_row = row
        #row = row + 1
    print('Zamykanie obiektu browser i zapisywanie pliku....')
    await browser.close()
    wb.save('prioritytire_scraper.xlsx')

async def main():
    print('Przygotowywanie adresów....')
    urls = [f'https://www.prioritytire.com/shop/?p={i+1}#/perpage:100' for i in range(200)]
    async with async_playwright() as playwright:
        await run(playwright, urls=urls)
    
s = time.time()
asyncio.run(main())
print("Asynchroniczne:", time.time() - s)